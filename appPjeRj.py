import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from selenium import webdriver as wb
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
from openpyxl import load_workbook
from selenium.webdriver.support import expected_conditions as EC
import pyautogui
import os
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')# Supondo que o seu código esteja em um arquivo chamado selenium_download.py

class Aplicativo(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Download de Processos PJE")
        self.geometry("400x300")
        self.arquivo_excel = None
        
        self.label_usuario = tk.Label(self, text="Usuário:")
        self.label_usuario.pack()
        self.entry_usuario = tk.Entry(self)
        self.entry_usuario.pack()
        
        self.label_senha = tk.Label(self, text="Senha:")
        self.label_senha.pack()
        self.entry_senha = tk.Entry(self, show="*")
        self.entry_senha.pack()
        
        self.label_arquivo = tk.Label(self, text="Arquivo Excel:")
        self.label_arquivo.pack()
        self.button_arquivo = tk.Button(self, text="Selecionar Arquivo", command=self.selecionar_arquivo)
        self.button_arquivo.pack()
        
        self.label_numero_processo = tk.Label(self, text="Número do Processo:")
        self.label_numero_processo.pack()
        self.entry_numero_processo = tk.Entry(self)
        self.entry_numero_processo.pack()
        
        self.button_iniciar = tk.Button(self, text="Iniciar Download", command=self.iniciar_download)
        self.button_iniciar.pack()

    def selecionar_arquivo(self):
        self.arquivo_excel = filedialog.askopenfilename(initialdir="/", title="Selecione o arquivo Excel",
                                                    filetypes=(("Excel Files", "*.xlsx"), ("All Files", "*.*")))
        
    def iniciar_download(self):
        usuario = self.entry_usuario.get()
        senha = self.entry_senha.get()
        numero_processo = self.entry_numero_processo.get()
        
        if not usuario or not senha:
            messagebox.showerror("Erro", "Por favor, preencha os campos de usuário e senha.")
        elif not (self.arquivo_excel or numero_processo):
            messagebox.showerror("Erro", "Selecione um arquivo Excel ou digite o número do processo.")
        else:
            if self.arquivo_excel:
                # Se um arquivo Excel foi selecionado, use o arquivo Excel
                self.download_files(usuario, senha, arquivo_excel=self.arquivo_excel)
            else:
                # Se não, use o número do processo digitado
                self.download_files(usuario, senha, numero_processo=numero_processo)

    def download_files_from_excel(self, usuario, senha, arquivo_excel):
        workbook = load_workbook(arquivo_excel)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            numero_processo = str(row[0])
            self.download_files(usuario, senha, numero_processo=numero_processo)
  
            
    def download_files(self, usuario, senha, arquivo_excel=None, numero_processo=None):
        # Configurar as opções do Chrome
        chrome_options = wb.ChromeOptions()
        
         # Obter o diretório inicial de downloads do sistema operacional
        diretorio_inicial = os.path.join(os.path.expanduser("~"), "Downloads")
        data_atual = datetime.now().strftime("%Y-%m-%d")
        diretorio_download = os.path.join(diretorio_inicial, data_atual)
        if not os.path.exists(diretorio_download):
            os.makedirs(diretorio_download)


        # Configurar a preferência de download
        prefs = {
            "download.default_directory": diretorio_download,
            "download.prompt_for_download": False,
        }
        chrome_options.add_experimental_option("prefs", prefs)
         # Inicialize o workbook antes de usar
        workbook = None
        if arquivo_excel:
            workbook = load_workbook(arquivo_excel)
            sheet = workbook.active
            coluna = sheet['A']
            
        # Instância do WebDriver
        driver = wb.Chrome(options=chrome_options)
        driver.implicitly_wait(30)
        driver.get('https://tjrj.pje.jus.br/1g/login.seam?loginComCertificado=false')
        time.sleep(1)

        # Especifique a coluna que deseja percorrer


        # Clicar no campo CPF/CNPJ*
        iframe = driver.find_element(By.XPATH, "//iframe[@id='ssoFrame']")
        driver.switch_to.frame(iframe)
        botaoCpf = driver.find_element(By.XPATH, "//*[@id='username']")
        botaoCpf.click()
        time.sleep(1)

        #DIGITAR O USUARIO
        botaoCpf.send_keys(usuario) # Conta da Juliana, Ex Processo: 0266495-94.2023.8.06.0001
        time.sleep(1)

        #CLICAR NA SENHA
        botaoPassword = driver.find_element(By.XPATH, "//*[@id='password']")
        botaoPassword.click()
        time.sleep(0.5)

        #DIGITA A SENHA 
        botaoPassword.send_keys(senha) # Conta da Juliana, Ex Processo: 0266495-94.2023.8.06.0001
        time.sleep(0.5)

        driver.find_element(By.XPATH, "//*[@id='kc-login']").click()
        time.sleep(0.5)


        #clica na barra menu
        driver.find_element(By.XPATH, "//*[@id='barraSuperiorPrincipal']/div/div[1]/ul/li/a").click()
        time.sleep(1)


        driver.find_element(By.XPATH, "//*[@id='menu']/div[2]/ul/li[2]/a").click()
        time.sleep(1)

        driver.find_element(By.XPATH, "//*[@id='menu']/div[2]/ul/li[2]/div/ul/li[4]").click()
        time.sleep(1)

        driver.find_element(By.XPATH, "//*[@id='menu']/div[2]/ul/li[2]/div/ul/li[4]/div/ul/li[1]").click()
        time.sleep(1)



        # Itere sobre as células da coluna e preencha o campo
        for cell in coluna:
            valor_limpo = limpar_valor(str(cell.value))

            #econtra o campo do numero do processo
            numeroProc = driver.find_element(By.XPATH, "//*[@id='fPP:numeroProcesso:numeroSequencial']")
            # Preencha o campo com o valor da célula atua
            numeroProc.send_keys(valor_limpo)
            time.sleep(2)

            #clica no botao pesquisar
            driver.find_element(By.XPATH, "//*[@id='fPP:searchProcessos']").click()
            time.sleep(2)


            #clica no botao pesquisar
            driver.find_element(By.XPATH, "/html/body/div[5]/div/div/div/div[2]/form/div[2]/div/table/tbody/tr/td[2]/a").click()
            time.sleep(5)

            #Clicar no alert
            pyautogui.press('enter')
            time.sleep(3)


            # Obtenha todas as alças de janelas abertas (janelas e abas)
            abas = driver.window_handles
            # Alterne para a nova janela (a segunda na lista de alças)
            driver.switch_to.window(abas[1])

            #clicar no dowloand
            driver.find_element(By.XPATH, "//*[@id='navbar:ajaxPanelAlerts']/ul[2]/li[5]/a").click()
            time.sleep(4)

            driver.find_element(By.XPATH, "//*[@id='navbar:downloadProcesso']").click()
            time.sleep(4)

            #Clicar no alert
            pyautogui.press('enter')
            time.sleep(4)

            # Fechar a aba[1]
            time.sleep(12)
            driver.close()

            # Alterne para a aba[0] (ou outra aba conforme necessário)
            abas = driver.window_handles
            driver.switch_to.window(abas[0])

            # Clicar para limpa 
            driver.find_element(By.XPATH, '//*[@id="fPP:clearButtonProcessos"]').click()
            time.sleep(5)

            #clica no campo processo
            time.sleep(3)
            driver.find_element(By.XPATH, "/html/body/div[5]/div/div/div/div[2]/form/div[1]/div/div/div[5]/div/div/div[2]/input[1]").click()
            time.sleep(8)
            
        messagebox.showinfo("Concluído", "Downloads concluídos com sucesso!")
        
def limpar_valor(valor):
    valor = valor.replace('-', '')  # Remove os "-"
    valor = valor.replace('8.19', '')  # Remove os "806"
    valor = valor.replace('.', '')  # Remove os "."
    return valor

if __name__ == "__main__":
    app = Aplicativo()
    app.mainloop()