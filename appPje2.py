import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from selenium import webdriver as wb
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
from openpyxl import load_workbook
import pyautogui
import os
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

class Aplicativo(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Download de Processos PJE")
        self.geometry("400x300")
        self.arquivo_excel = None
        self.downloads_concluidos = True  # Variável de controle
        
        self.label_usuario = tk.Label(self, text="Usuário:")
        self.label_usuario.pack()
        self.entry_usuario = tk.Entry(self)
        self.entry_usuario.pack()
        
        self.label_senha = tk.Label(self, text="Senha:")
        self.label_senha.pack()
        self.entry_senha = tk.Entry(self, show="*")
        self.entry_senha.pack()
        
        self.label_arquivo = tk.Label(self, text="Arquivo Excel:")
        self.label_arquivo.pack()
        self.button_arquivo = tk.Button(self, text="Selecionar Arquivo", command=self.selecionar_arquivo)
        self.button_arquivo.pack()
        
        self.label_caminho_arquivo = tk.Label(self, text="")
        self.label_caminho_arquivo.pack()
        
        self.label_numero_processo = tk.Label(self, text="Número do Processo:")
        self.label_numero_processo.pack()
        self.entry_numero_processo = tk.Entry(self)
        self.entry_numero_processo.pack()
        
        self.button_iniciar = tk.Button(self, text="Iniciar Download", command=self.iniciar_download)
        self.button_iniciar.pack()
        
    def selecionar_arquivo(self):
        self.arquivo_excel = filedialog.askopenfilename(initialdir="/", title="Selecione o arquivo Excel",
                                                    filetypes=(("Excel Files", "*.xlsx"), ("All Files", "*.*")))
        nome_arquivo = os.path.basename(self.arquivo_excel)
        # Atualiza o texto do rótulo com o caminho do arquivo selecionado
        self.label_caminho_arquivo.config(text=nome_arquivo)
    
    #função de download pelo arquivo selecionado 
    def iniciar_download(self):
        usuario = self.entry_usuario.get()
        senha = self.entry_senha.get()
        numero_processo = self.entry_numero_processo.get()
        
        if not usuario or not senha:
            messagebox.showerror("Erro", "Por favor, preencha os campos de usuário e senha.")
        elif not (self.arquivo_excel or numero_processo):
            messagebox.showerror("Erro", "Selecione um arquivo Excel ou digite o número do processo.")
        else:
            if self.arquivo_excel and not numero_processo:
                self.download_files_from_excel(usuario, senha, arquivo_excel=self.arquivo_excel)
            else:
                self.download_files(usuario, senha, numero_processo=numero_processo)
   
    #Função de reconhcer o processo individual
    def download_files_from_excel(self, usuario, senha, arquivo_excel):
        workbook = load_workbook(arquivo_excel)
        sheet = workbook.active
        row_number = 2  # Começando da linha 2, pois a linha 1 contém os cabeçalhos

        while True:
            numero_processo = str(sheet.cell(row=row_number, column=1).value)
            if not numero_processo:  # Se encontrar uma célula vazia, encerra o loop
                break
            if not self.download_files(usuario, senha, numero_processo=numero_processo):
                self.downloads_concluidos = False
            row_number += 1  # Avança para a próxima linha

        if self.downloads_concluidos:
            messagebox.showinfo("Concluído", "Todos os downloads foram concluídos com sucesso!")
        else:
            messagebox.showerror("Erro", "Alguns downloads falharam. Por favor, verifique o arquivo e tente novamente.")

    
            
    def download_files(self, usuario, senha, arquivo_excel=None, numero_processo=None):

        try:
            chrome_options = wb.ChromeOptions()
            diretorio_inicial = os.path.join(os.path.expanduser("~"), "Downloads")
            data_atual = datetime.now().strftime("%Y-%m-%d")
            diretorio_download = os.path.join(diretorio_inicial, data_atual)
            if not os.path.exists(diretorio_download):
                os.makedirs(diretorio_download)

            prefs = {
                "download.default_directory": diretorio_download,
                "download.prompt_for_download": False,
            }
            chrome_options.add_experimental_option("prefs", prefs)

            driver = wb.Chrome(options=chrome_options)
            driver.implicitly_wait(30)
            driver.get('https://tjrj.pje.jus.br/1g/login.seam?loginComCertificado=false')
            time.sleep(1)

            coluna = None
            if arquivo_excel:
                workbook = load_workbook(arquivo_excel)
                sheet = workbook.active
                coluna = sheet['A']
                
                

            
            iframe = driver.find_element(By.XPATH, "//iframe[@id='ssoFrame']")
            driver.switch_to.frame(iframe)
            botaoCpf = driver.find_element(By.XPATH, "//*[@id='username']")
            botaoCpf.click()
            time.sleep(1)

            botaoCpf.send_keys(usuario)
            time.sleep(1)

            botaoPassword = driver.find_element(By.XPATH, "//*[@id='password']")
            botaoPassword.click()
            time.sleep(0.5)

            botaoPassword.send_keys(senha)
            time.sleep(0.5)

            driver.find_element(By.XPATH, "//*[@id='kc-login']").click()
            time.sleep(0.5)

            driver.find_element(By.XPATH, "//*[@id='barraSuperiorPrincipal']/div/div[1]/ul/li/a").click()
            time.sleep(1)

            driver.find_element(By.XPATH, "//*[@id='menu']/div[2]/ul/li[2]/a").click()
            time.sleep(1)

            driver.find_element(By.XPATH, "//*[@id='menu']/div[2]/ul/li[2]/div/ul/li[4]").click()
            time.sleep(1)

            driver.find_element(By.XPATH, "//*[@id='menu']/div[2]/ul/li[2]/div/ul/li[4]/div/ul/li[1]").click()
            time.sleep(1)

            processos = []
            if numero_processo:
                processos.append(numero_processo)
            elif coluna:
                processos = [str(cell.value) for cell in coluna]
                
            processos_limpos = limpar_valor(processos)

            for valor_limpo in processos_limpos:
                    numeroProc = driver.find_element(By.XPATH, "//*[@id='fPP:numeroProcesso:numeroSequencial']")
                    numeroProc.send_keys(valor_limpo)
                    time.sleep(2)

                    driver.find_element(By.XPATH, "//*[@id='fPP:searchProcessos']").click()
                    time.sleep(2)

                    driver.find_element(By.XPATH, "/html/body/div[5]/div/div/div/div[2]/form/div[2]/div/table/tbody/tr/td[2]/a").click()
                    time.sleep(5)

                    pyautogui.press('enter')
                    time.sleep(3)

                    abas = driver.window_handles
                    driver.switch_to.window(abas[1])

                    driver.find_element(By.XPATH, "//*[@id='navbar:ajaxPanelAlerts']/ul[2]/li[5]/a").click()
                    time.sleep(4)

                    driver.find_element(By.XPATH, "//*[@id='navbar:downloadProcesso']").click()
                    time.sleep(4)

                    pyautogui.press('enter')
                    time.sleep(4)

                    time.sleep(12)
                    driver.close()

                    abas = driver.window_handles
                    driver.switch_to.window(abas[0])

                    driver.find_element(By.XPATH, '//*[@id="fPP:clearButtonProcessos"]').click()
                    time.sleep(5)

                    #clica no campo processo
                    time.sleep(3)
                    driver.find_element(By.XPATH, "/html/body/div[5]/div/div/div/div[2]/form/div[1]/div/div/div[5]/div/div/div[2]/input[1]").click()
                    time.sleep(8)
                    
            messagebox.showinfo("Concluído", "Downloads concluídos com sucesso!")
            return True

        except Exception as e:
            print(f"Erro: {str(e)}")
            return False
            
def limpar_valor(processos):
    processos_limpos = []
    for valor in processos:
        valor = valor.replace('-', '')  # Remove os "-"
        valor = valor.replace('8.19', '')  # Remove os "806"
        valor = valor.replace('.', '')  # Remove os "."
        processos_limpos.append(valor)
    return processos_limpos         


if __name__ == "__main__":
    app = Aplicativo()
    app.mainloop()
